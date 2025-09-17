# -*- coding: utf-8 -*-
# === VALIDACIONES IOS — Busca tabla BBDD_IOS_LAB en hoja BBDD_IOS; si no, usa la hoja completa ===

import io
import re
import unicodedata
from typing import Optional, Tuple, Dict, List

import pandas as pd
import streamlit as st

st.set_page_config(page_title="Validaciones IOS - BBDD_IOS", page_icon="🧪", layout="wide")
st.title("🕵️‍♂️ Validaciones IOS – BBDD_IOS")

# ====================== Dependencia opcional ======================
try:
    from openpyxl import load_workbook
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

# ====================== Sidebar ======================
with st.sidebar:
    st.header("⚙️ Configuración")
    hoja_fija = st.text_input("Hoja (fija)", value="BBDD_IOS")
    nombre_tabla = st.text_input("Tabla (fija)", value="BBDD_IOS_LAB")
    dayfirst = st.checkbox("Fechas DD/MM/AAAA", value=True)
    st.caption("Actívalo si tus fechas son del tipo 31/01/2025.")

uploaded = st.file_uploader("📁 Cargar Excel (.xlsx)", type=["xlsx"])

# ====================== Utilidades ======================
INVALID_SHEET_CHARS_RE = re.compile(r'[\[\]\:\*\?\/\\]')

def _sanitize_sheet_name(name: str, used: set) -> str:
    base = str(name) if name else "Hoja"
    base = base.replace("\n", " ").strip()
    base = INVALID_SHEET_CHARS_RE.sub(" ", base)
    if not base:
        base = "Hoja"
    base = base[:31]
    candidate = base
    i = 1
    while candidate in used:
        suffix = f"_{i}"
        maxlen = 31 - len(suffix)
        candidate = (base[:maxlen] + suffix) if maxlen > 0 else f"Hoja{suffix}"[:31]
        i += 1
    used.add(candidate)
    return candidate

def dataframe_to_excel_bytes_with_highlights(dfs: Dict[str, pd.DataFrame]) -> bytes:
    """
    Exporta varias hojas a Excel con:
      - Nombres de hoja saneados/únicos
      - Encabezado gris
      - Auto-filtro y panes congelados
      - Filas de hojas de ERRORES resaltadas en rojo (todas menos 'Resumen')
    """
    output = io.BytesIO()
    used = set()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        for name, d in dfs.items():
            df = d if isinstance(d, pd.DataFrame) else pd.DataFrame(d)
            safe_name = _sanitize_sheet_name(str(name), used)
            df.to_excel(writer, index=False, sheet_name=safe_name)
            wb  = writer.book
            ws  = writer.sheets[safe_name]
            nrows, ncols = df.shape

            # Formatos
            fmt_header = wb.add_format({'bold': True, 'bg_color': '#F2F2F2'})
            fmt_error  = wb.add_format({'bg_color': '#FFC7CE'})  # rojo suave
            # Encabezado
            ws.set_row(0, None, fmt_header)
            # Anchos y utilidades
            try:
                ws.set_column(0, ncols-1, 18)
            except Exception:
                pass
            if nrows > 0 and ncols > 0:
                ws.autofilter(0, 0, nrows, ncols-1)
            ws.freeze_panes(1, 0)

            # Resaltar en rojo todas las filas de hojas de errores (todas menos 'Resumen')
            if safe_name.lower() != "resumen" and nrows > 0 and ncols > 0:
                ws.conditional_format(
                    1, 0, nrows, ncols-1,
                    {'type': 'formula', 'criteria': '=TRUE', 'format': fmt_error}
                )
    return output.getvalue()

def strip_accents(s: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", str(s)) if not unicodedata.combining(ch))

def norm_header(s: str) -> str:
    # Normaliza encabezados: minúsculas, sin acentos, reemplaza espacios/_/.-/:/slash por espacio, colapsa espacios
    s = strip_accents(str(s)).lower().strip()
    s = re.sub(r"[\s_\.\-\/:]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s

def compact(s: str) -> str:
    return norm_header(s).replace(" ", "")

# ---- Aliases → canónico (fechas/resultado + contexto + extras + nuevas) ----
HEADER_ALIASES = {
    # Fechas / Resultado
    "fechasolicitud": "Fecha solicitud",
    "fechadesolicitud": "Fecha solicitud",
    "fecharecepcion": "Fecha solicitud",

    "fechatomademuestra": "Fecha toma de muestra",
    "fechadetomademuestra": "Fecha toma de muestra",
    "fechatoma": "Fecha toma de muestra",

    "fecharealizacion": "Fecha realización",
    "fechaderealizacion": "Fecha realización",
    "fecharesultado": "Fecha realización",

    "resultado": "Resultado",
    "resultadofinal": "Resultado",

    # Contexto base
    "pais": "País",
    "país": "País",
    "departamento": "Departamento",
    "depto": "Departamento",
    "municipio": "Municipio",
    "mun": "Municipio",
    "sitios": "SITIOS",
    "sitio": "SITIOS",
    "centro": "SITIOS",

    # Extra: Id/Expediente, Edad, Sexo
    "idpaciente": "Id paciente / No. Expediente",
    "iddelpaciente": "Id paciente / No. Expediente",
    "identificacionpaciente": "Id paciente / No. Expediente",
    "noexpediente": "Id paciente / No. Expediente",
    "nroexpediente": "Id paciente / No. Expediente",
    "numeroexpediente": "Id paciente / No. Expediente",
    "numerodeexpediente": "Id paciente / No. Expediente",
    "nexpediente": "Id paciente / No. Expediente",
    "expediente": "Id paciente / No. Expediente",
    "codigopaciente": "Id paciente / No. Expediente",

    "edad": "Edad",
    "anos": "Edad",
    "años": "Edad",
    "age": "Edad",

    "sexo": "Sexo",
    "genero": "Sexo",
    "género": "Sexo",
    "sex": "Sexo",

    # Nuevas columnas (ya existentes previamente)
    "seleccionelapruebaarealizar": "Seleccione la prueba a realizar",
    "seleccionepruebaarealizar": "Seleccione la prueba a realizar",
    "observaciones": "Observaciones",
    "observacion": "Observaciones",
    "observación": "Observaciones",

    # ===== Alias para VALIDACIONES NUEVAS =====
    # Laboratorio donde procesa la muestra
    "laboratoriodondeprocesalamuestra": "Laboratorio donde procesa la muestra",
    "laboratoriodondeprocesa": "Laboratorio donde procesa la muestra",
    "labdondeprocesalamuestra": "Laboratorio donde procesa la muestra",
    "labdondeprocesa": "Laboratorio donde procesa la muestra",
    "laboratorioprocesamuestra": "Laboratorio donde procesa la muestra",

    # Conteo CD4 (Cel/uL)
    "conteocd4celul": "Conteo CD4 (Cel/uL)",
    "conteocd4": "Conteo CD4 (Cel/uL)",
    "cd4conteo": "Conteo CD4 (Cel/uL)",
    "conteodecd4": "Conteo CD4 (Cel/uL)",
}

REQ_COLS_CANON = ["Fecha solicitud", "Fecha toma de muestra", "Fecha realización", "Resultado"]
CONTEXT_COLS_BASE = ["País", "Departamento", "Municipio", "SITIOS"]
EXTRA_COLS = ["Id paciente / No. Expediente", "Edad", "Sexo"]
# Observaciones irá **después de Resultado** en la salida; aquí solo nos aseguramos que exista
NEW_COLS = ["Seleccione la prueba a realizar", "Observaciones"]

# NUEVAS columnas a validar (no vacías)
VALIDATION_NEW_COLS = ["Laboratorio donde procesa la muestra", "Conteo CD4 (Cel/uL)"]

def to_datetime_safe(s, dayfirst=True):
    return pd.to_datetime(s, errors="coerce", dayfirst=dayfirst)

def is_empty_result(x) -> bool:
    if x is None:
        return True
    s = str(x).strip().lower()
    return s == "" or s in {"nan", "na", "none"}

def rename_to_canonical(df: pd.DataFrame) -> pd.DataFrame:
    """Renombra encabezados a canónicos cuando coinciden con alias; además, heurísticas para variantes."""
    newcols = {}
    for c in df.columns:
        key = compact(c)
        mapped = None
        if key in HEADER_ALIASES:
            mapped = HEADER_ALIASES[key]
        else:
            # Heurísticas por contención (p.ej., si viene con dos puntos, duplicado de texto, etc.)
            if "seleccionelapruebaarealizar" in key:
                mapped = "Seleccione la prueba a realizar"
            elif "observacion" in key or "observación" in key:
                mapped = "Observaciones"
            elif "laboratorio" in key and "procesa" in key and "muestra" in key:
                mapped = "Laboratorio donde procesa la muestra"
            elif ("conteo" in key and "cd4" in key) or ("cd4" in key and "cel" in key):
                mapped = "Conteo CD4 (Cel/uL)"
        if mapped:
            newcols[c] = mapped
    if newcols:
        df = df.rename(columns=newcols)
    return df

def ensure_columns(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    """Garantiza que existan las columnas (si faltan, las crea vacías)."""
    for col in cols:
        if col not in df.columns:
            df[col] = ""
    return df

def find_table_in_sheet(file_bytes: bytes, sheet_name: str, table_name: str) -> Optional[str]:
    if not OPENPYXL_OK:
        return None
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=False)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    tables = list(ws.tables.values()) if hasattr(ws, "tables") else list(getattr(ws, "_tables", []))
    for t in tables:
        tname = getattr(t, "name", None) or getattr(t, "displayName", None) or ""
        if tname and tname.strip().lower() == table_name.strip().lower():
            return t.ref
    return None

def read_table_range(file_bytes: bytes, sheet_name: str, ref: str) -> Tuple[pd.DataFrame, int]:
    start, end = ref.split(":")
    m1 = re.match(r"([A-Z]+)(\d+)$", start); m2 = re.match(r"([A-Z]+)(\d+)$", end)
    c1, r1 = m1.group(1), int(m1.group(2)); c2, r2 = m2.group(1), int(m2.group(2))
    usecols = f"{c1}:{c2}"; nrows = max(r2 - r1, 0)
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name,
                       header=r1-1, nrows=nrows, usecols=usecols,
                       engine="openpyxl", dtype=object)
    return df, r1

# ====================== Flujo principal ======================
if not uploaded:
    st.info("Sube el archivo para comenzar.")
    st.stop()

file_bytes = uploaded.read()
ref = find_table_in_sheet(file_bytes, hoja_fija, nombre_tabla)

if ref:
    st.success(f"Se encontró la tabla **{nombre_tabla}** en la hoja **{hoja_fija}** (rango {ref}).")
    df_raw, header_row_excel = read_table_range(file_bytes, hoja_fija, ref)
else:
    st.warning(f"No se encontró la tabla {nombre_tabla} en la hoja {hoja_fija}. Se usará la hoja completa.")
    df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=hoja_fija, engine="openpyxl", dtype=object)
    header_row_excel = 1

# Renombrar a canónicos (incluye contexto, extras y nuevas)
df = rename_to_canonical(df_raw.copy())

# Si faltan fechas/resultado, ofrecer mapeo manual mínimo
faltan = [c for c in REQ_COLS_CANON if c not in df.columns]
if faltan:
    st.subheader("🧭 Mapeo manual de columnas requeridas")
    st.caption("Selecciona manualmente si los encabezados difieren.")
    cols = list(df_raw.columns)

    def sel(label, pre):
        idx = cols.index(pre) + 1 if (pre in cols) else 0
        return st.selectbox(label, ["— Seleccionar —"] + cols, index=idx)

    def propose(target: str) -> Optional[str]:
        tgt = compact(target)
        for c in cols:
            if compact(c) == tgt:
                return c
        return None

    s_sol  = sel("Fecha solicitud", propose("Fecha solicitud") or "— Seleccionar —")
    s_toma = sel("Fecha toma de muestra", propose("Fecha toma de muestra") or "— Seleccionar —")
    s_real = sel("Fecha realización", propose("Fecha realización") or "— Seleccionar —")
    s_res  = sel("Resultado", propose("Resultado") or "— Seleccionar —")

    if any(x == "— Seleccionar —" for x in [s_sol, s_toma, s_real, s_res]):
        st.error("Faltan columnas por seleccionar. Completa los 4 campos.")
        st.stop()

    df = df_raw.rename(columns={
        s_sol: "Fecha solicitud",
        s_toma: "Fecha toma de muestra",
        s_real: "Fecha realización",
        s_res: "Resultado",
    })

# Asegurar columnas de contexto, extras, nuevas y las de VALIDACIÓN NUEVA (antes del filtrado)
pre_result_context = CONTEXT_COLS_BASE + ["Seleccione la prueba a realizar"] + EXTRA_COLS
df = ensure_columns(df, pre_result_context + ["Observaciones"] + VALIDATION_NEW_COLS)

# Añadir número de fila real de Excel
df["Fila (Excel)"] = header_row_excel + 1 + df.index

# Parsear fechas
df["Fecha solicitud"] = to_datetime_safe(df["Fecha solicitud"], dayfirst=dayfirst)
df["Fecha toma de muestra"] = to_datetime_safe(df["Fecha toma de muestra"], dayfirst=dayfirst)
df["Fecha realización"] = to_datetime_safe(df["Fecha realización"], dayfirst=dayfirst)

total = len(df)

# ====================== Validaciones ======================
v_solicitud = df[df["Fecha solicitud"].isna()]
v_toma      = df[df["Fecha toma de muestra"].isna()]
v_realiz    = df[df["Fecha realización"].isna()]
v_result    = df[df["Resultado"].apply(is_empty_result)]

# NUEVAS: no vacío en "Laboratorio..." y "Conteo CD4..."
v_lab_vacio = df[df["Laboratorio donde procesa la muestra"].apply(is_empty_result)]
v_cd4_vacio = df[df["Conteo CD4 (Cel/uL)"].apply(is_empty_result)]

err_toma_lt_sol = df[
    df["Fecha toma de muestra"].notna()
    & df["Fecha solicitud"].notna()
    & (df["Fecha toma de muestra"] < df["Fecha solicitud"])
]
err_real_lt_toma = df[
    df["Fecha realización"].notna()
    & df["Fecha toma de muestra"].notna()
    & (df["Fecha realización"] < df["Fecha toma de muestra"])
]
err_real_lt_sol = df[
    df["Fecha realización"].notna()
    & df["Fecha solicitud"].notna()
    & (df["Fecha realización"] < df["Fecha solicitud"])
]

# ============ ORDEN FINAL PARA LAS SALIDAS ============
# Observaciones **después de Resultado**
cols_show = (
    pre_result_context
    + ["Laboratorio donde procesa la muestra", "Conteo CD4 (Cel/uL)"]
    + ["Fila (Excel)", "Fecha solicitud", "Fecha toma de muestra", "Fecha realización", "Resultado", "Observaciones"]
)

def with_missing_cols(dsub: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    out = dsub.copy()
    for c in cols:
        if c not in out.columns:
            out[c] = ""
    return out[cols]

tablas = {
    "Vacías - Fecha solicitud": with_missing_cols(v_solicitud, cols_show),
    "Vacías - Fecha toma de muestra": with_missing_cols(v_toma, cols_show),
    "Vacías - Fecha realización": with_missing_cols(v_realiz, cols_show),
    "Resultado vacío": with_missing_cols(v_result, cols_show),

    # NUEVAS
    "Vacías - Lab procesa muestra": with_missing_cols(v_lab_vacio, cols_show),
    "Vacías - Conteo CD4 (Cel/uL)": with_missing_cols(v_cd4_vacio, cols_show),

    "Orden: toma < solicitud": with_missing_cols(err_toma_lt_sol, cols_show),
    "Orden: realiz < toma": with_missing_cols(err_real_lt_toma, cols_show),
    "Orden: realiz < solicitud": with_missing_cols(err_real_lt_sol, cols_show),
}

resumen = pd.DataFrame([
    {"Error": "Fecha solicitud vacía", "Conteo": len(v_solicitud)},
    {"Error": "Fecha toma de muestra vacía", "Conteo": len(v_toma)},
    {"Error": "Fecha realización vacía", "Conteo": len(v_realiz)},
    {"Error": "Resultado vacío", "Conteo": len(v_result)},

    # NUEVAS en resumen
    {"Error": "Laboratorio procesa muestra vacío", "Conteo": len(v_lab_vacio)},
    {"Error": "Conteo CD4 (Cel/uL) vacío", "Conteo": len(v_cd4_vacio)},

    {"Error": "toma < solicitud", "Conteo": len(err_toma_lt_sol)},
    {"Error": "realización < toma", "Conteo": len(err_real_lt_toma)},
    {"Error": "realización < solicitud", "Conteo": len(err_real_lt_sol)},
])
resumen["% del total"] = (resumen["Conteo"] / total * 100).round(2) if total else 0.0

# ====================== UI ======================
# ====================== UI ======================
st.subheader("📊 Resumen de validaciones")

# Fila 0: totales y agregados (se mantienen)
total_fechas_vacias = int(
    resumen[resumen["Error"].isin([
        "Fecha solicitud vacía",
        "Fecha toma de muestra vacía",
        "Fecha realización vacía",
    ])]["Conteo"].sum()
)
total_reglas_orden = int(
    resumen[resumen["Error"].isin([
        "toma < solicitud",
        "realización < toma",
        "realización < solicitud",
    ])]["Conteo"].sum()
)
resultado_vacio = int(
    resumen.loc[resumen["Error"] == "Resultado vacío", "Conteo"].sum()
)

c1, c2, c3, c4 = st.columns(4)
c1.metric("Registros totales", f"{total:,}")
#c2.metric("Fechas vacías (suma)", f"{total_fechas_vacias:,}")
#c3.metric("Resultado vacío", f"{resultado_vacio:,}")
#c4.metric("Reglas de orden incumplidas", f"{total_reglas_orden:,}")

# Filas siguientes: tarjetas para TODAS las validaciones (dinámico)
res_sorted = resumen.sort_values("Conteo", ascending=False).reset_index(drop=True)
per_row = 4  # cambia a 3 o 5 si prefieres otro ancho
for i in range(0, len(res_sorted), per_row):
    cols = st.columns(per_row)
    chunk = res_sorted.iloc[i:i+per_row]
    for col, (_, row) in zip(cols, chunk.iterrows()):
        # Muestra el % del total en el "delta"
        col.metric(row["Error"], f"{int(row['Conteo']):,}", f"{row['% del total']}%")

# Mantén la tabla resumen debajo
st.dataframe(res_sorted, use_container_width=True)


st.subheader("🧾 Detalle por tipo de error")
for nombre, d in tablas.items():
    with st.expander(f"{nombre} ({len(d)})", expanded=False):
        if d.empty:
            st.info("Sin registros.")
        else:
            st.dataframe(d, use_container_width=True)

# ====================== Descargas (con resaltado en rojo) ======================
st.subheader("⬇️ Descargar resultados")
to_export = {"Resumen": resumen}
to_export.update(tablas)
xlsx_bytes = dataframe_to_excel_bytes_with_highlights(to_export)
st.download_button(
    "Descargar Excel (resumen + errores resaltados)",
    data=xlsx_bytes,
    file_name="validaciones_bbdd_ios_lab.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# (opcional) Válidos sin violaciones de orden ni fechas vacías
invalid_idx = (
    set(v_solicitud.index) | set(v_toma.index) | set(v_realiz.index) |
    set(err_toma_lt_sol.index) | set(err_real_lt_toma.index) | set(err_real_lt_sol.index)
)
df_validos = df[~df.index.isin(invalid_idx)].copy()
df_validos = with_missing_cols(df_validos, cols_show)  # asegurar columnas y orden final (Obs. tras Resultado)
csv_validos = df_validos.to_csv(index=False).encode("utf-8-sig")
st.download_button("Descargar CSV de registros válidos", data=csv_validos,
                   file_name="registros_validos.csv", mime="text/csv")










